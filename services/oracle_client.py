from __future__ import annotations

from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple


def _format_date(value):
    return value.strftime("%d/%m/%Y") if value else None


class OracleClient:
    """Client Oracle centralisé (initialisation + requêtes Assurés)."""

    def __init__(
        self,
        dsn: str,
        instant_client_dir: Optional[Path] = None,
        oracle_config_dir: Optional[Path] = None,
    ) -> None:
        self.dsn = dsn
        self.instant_client_dir = instant_client_dir
        self.oracle_config_dir = oracle_config_dir
        self.oracle_client_init_error: Optional[str] = None
        self.instant_client_dir_used: Optional[str] = None
        self.oracle_config_dir_used: Optional[str] = None
        self._init_oracle_client()

    def _format_oracle_error(self, exc: BaseException) -> str:
        msg = str(exc)
        if "ORA-28041" in msg:
            extra = (
                "ORA-28041 est souvent lie a une negotiation d'authentification (NTS/SSO) cote client.\n"
                "Solution recommandee: fournir un sqlnet.ora minimal (SQLNET.AUTHENTICATION_SERVICES=(NONE))\n"
                "et forcer son utilisation via TNS_ADMIN / config_dir (Instant Client en mode THICK)."
            )
            if self.oracle_config_dir_used or self.oracle_config_dir:
                extra = f"{extra}\n\noracle_config_dir: {self.oracle_config_dir_used or str(self.oracle_config_dir)}"

            try:
                d = self.diagnostics()
                mode = (d.get("mode") or "unknown").upper()
                v = d.get("client_version")
                dsn_hint = d.get("dsn")
                extra = f"{extra}\nmode: {mode}{' - client ' + str(v) if v else ''}\ndsn: {dsn_hint}"
            except Exception:
                pass
            return f"{msg}\n\n{extra}"
        if "DPY-3010" in msg:
            extra = (
                "Ce serveur Oracle n'est pas supporte par python-oracledb en mode THIN.\n"
                "Solution: utiliser le mode THICK avec Oracle Instant Client (installe ou embarque avec l'application)."
            )
            if self.oracle_client_init_error:
                extra = f"{extra}\n\nErreur Oracle Instant Client:\n{self.oracle_client_init_error}"
            return f"{msg}\n\n{extra}"
        return msg

    def _write_oracle_error_log(self, exc: BaseException) -> None:
        """
        Write a local log file (next to EXE when frozen) for troubleshooting issues on other PCs.
        Never logs credentials.
        """
        try:
            import sys

            if getattr(sys, "frozen", False):
                root = Path(sys.executable).parent
            else:
                root = Path(__file__).resolve().parents[1]

            log_path = root / "oracle_error.log"
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            diag = {}
            try:
                diag = self.diagnostics()
            except Exception:
                diag = {}

            parts = [
                f"[{ts}] Oracle connection error",
                f"error: {type(exc).__name__}: {exc}",
                f"mode: {diag.get('mode')}",
                f"client_version: {diag.get('client_version')}",
                f"instant_client_dir: {diag.get('instant_client_dir')}",
                f"oracle_config_dir: {diag.get('oracle_config_dir')}",
                f"dsn: {diag.get('dsn')}",
                "",
            ]
            log_path.write_text("\n".join(parts), encoding="utf-8")
        except Exception:
            # Best-effort only.
            return

    def _init_oracle_client(self) -> None:
        """Initialise l'instant client Oracle si présent (optionnel)."""
        try:
            import oracledb
        except Exception:
            return

        if self.instant_client_dir and self.instant_client_dir.exists():
            try:
                config_dir = None
                if self.oracle_config_dir and self.oracle_config_dir.exists():
                    config_dir = str(self.oracle_config_dir)

                # config_dir is crucial for sqlnet.ora/tnsnames.ora resolution in frozen builds.
                oracledb.init_oracle_client(lib_dir=str(self.instant_client_dir), config_dir=config_dir)
                self.instant_client_dir_used = str(self.instant_client_dir)
                self.oracle_config_dir_used = config_dir
            except Exception as exc:  # noqa: BLE001
                # Keep running, but remember the error to help troubleshooting.
                self.oracle_client_init_error = str(exc)

    def diagnostics(self) -> Dict[str, object]:
        """Return diagnostics about the Oracle driver mode for UI/troubleshooting."""
        try:
            import oracledb
        except Exception as exc:
            return {"ok": "false", "mode": "unknown", "error": f"Module oracledb indisponible : {exc}"}

        mode = "thin" if getattr(oracledb, "is_thin", lambda: True)() else "thick"
        client_version = None
        if mode == "thick":
            try:
                v = oracledb.clientversion()
                client_version = ".".join(str(x) for x in v)
            except Exception:
                client_version = None

        # Don't leak full DSN details; provide a short hint.
        dsn_hint = self.dsn
        if len(dsn_hint) > 120:
            dsn_hint = dsn_hint[:117] + "..."

        return {
            "ok": "true",
            "mode": mode,
            "client_version": client_version,
            "instant_client_dir": self.instant_client_dir_used or (str(self.instant_client_dir) if self.instant_client_dir else None),
            "oracle_config_dir": self.oracle_config_dir_used or (str(self.oracle_config_dir) if self.oracle_config_dir else None),
            "instant_client_error": self.oracle_client_init_error,
            "dsn": dsn_hint,
        }

    def test_connection(self, username: str, password: str) -> Optional[str]:
        """Retourne None si OK, sinon le message d'erreur."""
        try:
            import oracledb
        except Exception as exc:
            return f"Module oracledb indisponible : {exc}"

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute("SELECT 1 FROM dual")
                    cursor.fetchone()
            return None
        except Exception as exc:  # noqa: BLE001
            self._write_oracle_error_log(exc)
            return self._format_oracle_error(exc)

    def query_user_login(self, username: str, password: str, identifier: str, user_password: str) -> Dict[str, object]:
        """Vérifie un utilisateur applicatif par identifiant + mot de passe."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": None, "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                usr.user_id,
                usr.prenom,
                usr.nom,
                usr.email,
                usr.mot_de_passe,
                usr.service,
                usr.identifiant
            FROM USER_IDENT usr
            WHERE TRIM(UPPER(usr.identifiant)) = TRIM(UPPER(:identifier))
        """
        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"identifier": identifier})
                    row = cursor.fetchone()
        except Exception as exc:  # noqa: BLE001
            return {"data": None, "error": self._format_oracle_error(exc)}

        if not row:
            return {"data": None, "error": None}

        stored_pwd = (row[4] or "").strip()
        provided_pwd = (user_password or "").strip()
        if stored_pwd != provided_pwd:
            return {"data": None, "error": None}

        data = {
            "user_id": row[0],
            "prenoms": row[1],
            "nom": row[2],
            "email": row[3],
            "service": row[5],
            "identifiant": row[6],
        }
        return {"data": data, "error": None}

    def query_assures(
        self,
        username: str,
        password: str,
        nir_value: Optional[str],
        nom_value: Optional[str],
        prenom_value: Optional[str],
        order_by: str = "nir",
    ) -> Dict[str, object]:
        """Exécute la requête assurés par filtres (préfixe), insensible à la casse."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        nir_pattern = f"{nir_value}%" if nir_value else None
        nom_pattern = f"{nom_value}%" if nom_value else None
        prenom_pattern = f"{prenom_value}%" if prenom_value else None

        order_field = {
            "nir": "a.as_nni",
            "nom": "a.as_nompat",
            "prenom": "a.as_prenoms",
        }.get(order_by, "a.as_nni")

        sql = """
            SELECT
                a.as_nni,
                a.as_nompat,
                a.as_prenoms,
                a.as_dtnais,
                cv.cv_lib,
                a.as_nomusuel,
                a.as_dtrniam,
                a.as_naiscp,
                a.as_codcomnais,
                a.as_naiscom,
                p0.py_lib,
                tn.tn_lib,
                p1.py_lib,
                p2.py_lib,
                a.as_dtnatur,
                a.as_dtvierel,
                a.as_dtcesact,
                a.as_dtfin_visa,
                a.as_dtmaj,
                COALESCE(
                    MAX(CASE WHEN ac.acst_id = 1 THEN ac.accl_id END),
                    a.ascocl_id
                ) AS collectivite_maladie,
                COALESCE(
                    MAX(CASE WHEN ac.acst_id = 2 THEN ac.accl_id END),
                    a.asco_id
                ) AS collectivite_vieillesse
            FROM AT_AS#ASSURE a
            LEFT JOIN AT_CV#CIVILITE cv ON a.ascv_id = cv.cv_id
            LEFT JOIN AT_PY#PAYS p0 ON a.aspy_nais_id = p0.py_id
            LEFT JOIN AT_TN#TYPE_NATIONAL tn ON a.astn_id = tn.tn_id
            LEFT JOIN AT_PY#PAYS p1 ON a.aspy_natpays1_id = p1.py_id
            LEFT JOIN AT_PY#PAYS p2 ON a.aspy_natpays2_id = p2.py_id
            LEFT JOIN AT_AC#ASS_COL ac
              ON ac.acas_id = a.as_id
             AND ac.ac_dtefdeb <= SYSDATE
             AND (ac.ac_dteffin IS NULL OR ac.ac_dteffin = TO_DATE('31123999', 'DDMMYYYY'))
            WHERE ( :nir_pattern IS NULL OR UPPER(a.as_nni) LIKE UPPER(:nir_pattern) )
              AND (
                    :nom_pattern IS NULL
                    OR UPPER(a.as_nompat) LIKE UPPER(:nom_pattern)
                    OR ( :prenom_pattern IS NULL AND UPPER(a.as_prenoms) LIKE UPPER(:nom_pattern) )
                  )
              AND ( :prenom_pattern IS NULL OR UPPER(a.as_prenoms) LIKE UPPER(:prenom_pattern) )
            GROUP BY
                a.as_nni,
                a.as_nompat,
                a.as_prenoms,
                a.as_dtnais,
                cv.cv_lib,
                a.as_nomusuel,
                a.as_dtrniam,
                a.as_naiscp,
                a.as_codcomnais,
                a.as_naiscom,
                p0.py_lib,
                tn.tn_lib,
                p1.py_lib,
                p2.py_lib,
                a.as_dtnatur,
                a.as_dtvierel,
                a.as_dtcesact,
                a.as_dtfin_visa,
                a.as_dtmaj,
                a.ascocl_id,
                a.asco_id
        """

        sql = f"{sql} ORDER BY {order_field}"

        try:
            params = {
                "nir_pattern": nir_pattern,
                "nom_pattern": nom_pattern,
                "prenom_pattern": prenom_pattern,
            }
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, params)
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append(
                {
                    "nir": row[0],
                    "nom_usage": row[1],
                    "prenom_usage": row[2],
                    "date_naissance": _format_date(row[3]),
                    "civilite": row[4],
                    "nom_usuel": row[5],
                    "date_certif_rniam": _format_date(row[6]),
                    "code_postal": row[7],
                    "code_commune_naissance": row[8],
                    "commune": row[9],
                    "pays_naissance": row[10],
                    "type_nationalite": row[11],
                    "pays1": row[12],
                    "pays2": row[13],
                    "date_naturalisation": _format_date(row[14]),
                    "date_entree_vie_religieuse": _format_date(row[15]),
                    "date_cessation_vie_religieuse": _format_date(row[16]),
                    "date_fin_visa": _format_date(row[17]),
                    "date_maj": _format_date(row[18]),
                    "collectivite_maladie": row[19],
                    "collectivite_vieillesse": row[20],
                    "sexe": None,
                    "email": None,
                    "pays_nationalite": None,
                    "date_deces": None,
                    "type_contrat_individu": None,
                    "numero_adherent": None,
                    "raison_sociale": None,
                    "date_effet_contrat": None,
                    "date_conditions_contrat": None,
                }
            )
        return {"data": data, "error": None}

    def query_assure_situations_maladie(
        self,
        username: str,
        password: str,
        nir: str,
    ) -> Dict[str, object]:
        """Retourne les situations maladie d'un assure (par NIR)."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                sm.smsa_id,
                sm.smns_id,
                sm.smns_id2,
                sm.sm_dtsitnat,
                sm.sm_dtcond,
                sm.sm_dtdecl,
                sm.sm_dtefdeb,
                sm.sm_dtnot,
                sm.sm_dtmaj
            FROM AT_AS#ASSURE a
            LEFT JOIN AT_SM#ASS_SIT_CAM sm
              ON sm.smas_id = a.as_id
            WHERE a.as_nni = :nir
            ORDER BY sm.sm_dtefdeb DESC
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"nir": nir})
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append(
                {
                    "code_situation": row[0],
                    "code_nature1": row[1],
                    "code_nature2": row[2],
                    "date_nature2": _format_date(row[3]),
                    "date_conditions": _format_date(row[4]),
                    "date_declaration": _format_date(row[5]),
                    "date_effet": _format_date(row[6]),
                    "cristallisation": None,
                    "date_maj_situation": _format_date(row[7]),
                    "date_maj": _format_date(row[8]),
                }
            )
        return {"data": data, "error": None}

    def query_assure_situation_maladie_current(
        self,
        username: str,
        password: str,
        nir: str,
    ) -> Dict[str, object]:
        """Retourne la situation maladie en cours d'un assure (par NIR)."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                sm.smsa_id,
                sm.smns_id,
                sm.smns_id2,
                sm.sm_dtsitnat,
                sm.sm_dtcond,
                sm.sm_dtdecl,
                sm.sm_dtefdeb,
                sm.sm_dtnot,
                sm.sm_dtmaj
            FROM AT_AS#ASSURE a
            JOIN AT_SM#ASS_SIT_CAM sm
              ON sm.smas_id = a.as_id
            WHERE a.as_nni = :nir
              AND sm.sm_dteffin = TO_DATE('31123999', 'DDMMYYYY')
            ORDER BY sm.sm_dtefdeb DESC
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"nir": nir})
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append(
                {
                    "code_situation": row[0],
                    "code_nature1": row[1],
                    "code_nature2": row[2],
                    "date_nature2": _format_date(row[3]),
                    "date_conditions": _format_date(row[4]),
                    "date_declaration": _format_date(row[5]),
                    "date_effet": _format_date(row[6]),
                    "date_maj_situation": _format_date(row[7]),
                    "date_maj": _format_date(row[8]),
                }
            )
        return {"data": data, "error": None}

    def query_assure_situation_vieillesse_current(
        self,
        username: str,
        password: str,
        nir: str,
    ) -> Dict[str, object]:
        """Retourne la situation vieillesse en cours d'un assure (par NIR)."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                sv.svsa_id,
                sv.svns_id,
                sv.svns_id2,
                NULL,
                sv.sv_dtcond,
                sv.sv_dtdecl,
                sv.sv_dtefdeb,
                sv.sv_dtnot,
                sv.sv_dtmaj
            FROM AT_AS#ASSURE a
            JOIN AT_SV#ASS_SIT_VIC sv
              ON sv.svas_id = a.as_id
            WHERE a.as_nni = :nir
              AND sv.sv_dteffin = TO_DATE('31123999', 'DDMMYYYY')
            ORDER BY sv.sv_dtefdeb DESC
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"nir": nir})
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append(
                {
                    "code_situation": row[0],
                    "code_nature1": row[1],
                    "code_nature2": row[2],
                    "date_nature2": _format_date(row[3]),
                    "date_conditions": _format_date(row[4]),
                    "date_declaration": _format_date(row[5]),
                    "date_effet": _format_date(row[6]),
                    "date_maj_situation": _format_date(row[7]),
                    "date_maj": _format_date(row[8]),
                }
            )
        return {"data": data, "error": None}

    def query_assure_historique_situation_maladie(
        self,
        username: str,
        password: str,
        nir: str,
    ) -> Dict[str, object]:
        """Retourne l'historique des situations maladie (par NIR)."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                sm.smsa_id,
                sm.smns_id,
                sm.smns_id2,
                sm.sm_dtsitnat,
                sm.sm_dtcond,
                sm.sm_dtdecl,
                sm.sm_dtefdeb,
                sm.sm_dtnot,
                sm.sm_dtmaj
            FROM AT_AS#ASSURE a
            LEFT JOIN AT_SM#ASS_SIT_CAM sm
              ON sm.smas_id = a.as_id
            WHERE a.as_nni = :nir
            ORDER BY sm.sm_dtefdeb DESC
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"nir": nir})
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append(
                {
                    "code_situation": row[0],
                    "code_nature1": row[1],
                    "code_nature2": row[2],
                    "date_nature2": _format_date(row[3]),
                    "date_conditions": _format_date(row[4]),
                    "date_declaration": _format_date(row[5]),
                    "date_effet": _format_date(row[6]),
                    "date_maj_situation": _format_date(row[7]),
                    "date_maj": _format_date(row[8]),
                }
            )
        return {"data": data, "error": None}

    def query_assure_collectivites_maladie(
        self,
        username: str,
        password: str,
        nir: str,
    ) -> Dict[str, object]:
        """Retourne l'historique des collectivites maladie (par NIR)."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                ac.accl_id,
                ac.ac_dtefdeb,
                ac.ac_dtmaj,
                cs.cs_lib
            FROM AT_AC#ASS_COL ac
            JOIN AT_AS#ASSURE a
              ON ac.acas_id = a.as_id
            LEFT JOIN AT_HC#HIS_SIT_COL hc
              ON hc.hccl_id = ac.accl_id
             AND hc.hc_dtfin = TO_DATE('31123999', 'DDMMYYYY')
            LEFT JOIN AT_CS#LIB_SIT_COL cs
              ON cs.cs_id = hc.hccs_id
            WHERE a.as_nni = :nir
              AND ac.acst_id = '1'
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"nir": nir})
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append(
                {
                    "collectivite": row[0],
                    "date_effet": _format_date(row[1]),
                    "date_maj": _format_date(row[2]),
                    "etat_actuel": row[3],
                }
            )
        return {"data": data, "error": None}

    def query_assure_renouvellement_regime_particulier_maladie(
        self,
        username: str,
        password: str,
        nir: str,
    ) -> Dict[str, object]:
        """Retourne les dates de renouvellement du regime particulier (maladie) pour un NIR."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                sm.sm_dtef_rnv,
                sm.sm_dtenv_rnv,
                sm.sm_dtret_rnv
            FROM AT_AS#ASSURE a
            JOIN AT_SM#ASS_SIT_CAM sm
              ON sm.smas_id = a.as_id
            WHERE a.as_nni = :nir
            ORDER BY sm.sm_dtefdeb DESC
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"nir": nir})
                    row = cursor.fetchone()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        if not row:
            return {"data": [], "error": None}

        data = [
            {
                "date_effet_dernier_renouvellement": _format_date(row[0]),
                "date_envoi_renouvellement": _format_date(row[1]),
                "date_retour_renouvellement": _format_date(row[2]),
            }
        ]
        return {"data": data, "error": None}

    def query_assure_historique_situation_vieillesse(
        self,
        username: str,
        password: str,
        nir: str,
    ) -> Dict[str, object]:
        """Retourne l'historique des situations vieillesse (par NIR)."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                sv.svsa_id,
                sv.svns_id,
                sv.svns_id2,
                sv.sv_dtsitnat,
                sv.sv_dtcond,
                sv.sv_dtdecl,
                sv.sv_dtefdeb,
                sv.sv_dtnot,
                sv.sv_dtmaj
            FROM AT_AS#ASSURE a
            LEFT JOIN AT_SV#ASS_SIT_VIC sv
              ON sv.svas_id = a.as_id
            WHERE a.as_nni = :nir
            ORDER BY sv.sv_dtefdeb DESC
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"nir": nir})
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append(
                {
                    "code_situation": row[0],
                    "code_nature1": row[1],
                    "code_nature2": row[2],
                    "date_nature2": _format_date(row[3]),
                    "date_conditions": _format_date(row[4]),
                    "date_declaration": _format_date(row[5]),
                    "date_effet": _format_date(row[6]),
                    "date_maj_situation": _format_date(row[7]),
                    "date_maj": _format_date(row[8]),
                }
            )
        return {"data": data, "error": None}

    def query_assure_collectivites_vieillesse(
        self,
        username: str,
        password: str,
        nir: str,
    ) -> Dict[str, object]:
        """Retourne l'historique des collectivites vieillesse (par NIR)."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                ac.accl_id,
                ac.ac_dtefdeb,
                ac.ac_dtmaj,
                cs.cs_lib
            FROM AT_AC#ASS_COL ac
            JOIN AT_AS#ASSURE a
              ON ac.acas_id = a.as_id
            LEFT JOIN AT_HC#HIS_SIT_COL hc
              ON hc.hccl_id = ac.accl_id
             AND hc.hc_dtfin = TO_DATE('31123999', 'DDMMYYYY')
            LEFT JOIN AT_CS#LIB_SIT_COL cs
              ON cs.cs_id = hc.hccs_id
            WHERE a.as_nni = :nir
              AND ac.acst_id = '2'
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"nir": nir})
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append(
                {
                    "collectivite": row[0],
                    "date_effet": _format_date(row[1]),
                    "date_maj": _format_date(row[2]),
                    "etat_actuel": row[3],
                }
            )
        return {"data": data, "error": None}

    def query_assure_historique_situation_rco(
        self,
        username: str,
        password: str,
        nir: str,
    ) -> Dict[str, object]:
        """Retourne l'historique des situations RCO (par NIR)."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                sr.srcsa_id,
                sr.src_dtcond,
                sr.src_dtdecl,
                sr.src_dtefdeb,
                sr.src_dtnot,
                sr.src_dtmaj
            FROM AT_AS#ASSURE a
            LEFT JOIN AT_SRC#ASS_SIT_RCO sr
              ON sr.srcas_id = a.as_id
            WHERE a.as_nni = :nir
            ORDER BY sr.src_dtefdeb DESC
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"nir": nir})
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append(
                {
                    "code_situation": row[0],
                    "date_conditions": _format_date(row[1]),
                    "date_declaration": _format_date(row[2]),
                    "date_effet": _format_date(row[3]),
                    "date_maj_situation": _format_date(row[4]),
                    "date_maj": _format_date(row[5]),
                }
            )
        return {"data": data, "error": None}

    def query_assure_collectivites_rco(
        self,
        username: str,
        password: str,
        nir: str,
    ) -> Dict[str, object]:
        """
        Retourne l'historique des collectivites RCO (par NIR).

        Note: selon le besoin fourni, ACST_ID = '2'.
        """
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                ac.accl_id,
                ac.ac_dtefdeb,
                ac.ac_dtmaj,
                cs.cs_lib
            FROM AT_AC#ASS_COL ac
            JOIN AT_AS#ASSURE a
              ON ac.acas_id = a.as_id
            LEFT JOIN AT_HC#HIS_SIT_COL hc
              ON hc.hccl_id = ac.accl_id
             AND hc.hc_dtfin = TO_DATE('31123999', 'DDMMYYYY')
            LEFT JOIN AT_CS#LIB_SIT_COL cs
              ON cs.cs_id = hc.hccs_id
            WHERE a.as_nni = :nir
              AND ac.acst_id = '2'
            ORDER BY ac.ac_dtefdeb DESC
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"nir": nir})
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append(
                {
                    "collectivite": row[0],
                    "date_effet": _format_date(row[1]),
                    "date_maj": _format_date(row[2]),
                    "etat_actuel": row[3],
                }
            )
        return {"data": data, "error": None}

    def query_assure_collectivites_csg(
        self,
        username: str,
        password: str,
        nir: str,
    ) -> Dict[str, object]:
        """Retourne l'historique des collectivites maladie pour l'onglet CSG (par NIR)."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                ac.accl_id,
                ac.ac_dtefdeb,
                ac.ac_dtmaj,
                cs.cs_lib
            FROM AT_AC#ASS_COL ac
            JOIN AT_AS#ASSURE a
              ON ac.acas_id = a.as_id
            LEFT JOIN AT_HC#HIS_SIT_COL hc
              ON hc.hccl_id = ac.accl_id
             AND hc.hc_dtfin = TO_DATE('31123999', 'DDMMYYYY')
            LEFT JOIN AT_CS#LIB_SIT_COL cs
              ON cs.cs_id = hc.hccs_id
            WHERE a.as_nni = :nir
              AND ac.acst_id = '1'
            ORDER BY ac.ac_dtefdeb DESC
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"nir": nir})
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append(
                {
                    "collectivite": row[0],
                    "date_effet": _format_date(row[1]),
                    "date_maj": _format_date(row[2]),
                    "etat_actuel": row[3],
                }
            )
        return {"data": data, "error": None}

    def query_assure_adresse(
        self,
        username: str,
        password: str,
        nir: str,
    ) -> Dict[str, object]:
        """Retourne l'adresse et les coordonnees de l'assure (par NIR)."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": None, "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                aa.aa_adr1,
                aa.aa_adr2,
                aa.aa_adr3,
                aa.aa_adr4,
                aa.aa_adrcp,
                aa.aa_adrvil,
                py.py_lib,
                aa.aa_email,
                aa.aa_tel1,
                aa.aa_tel2,
                aa.aa_tel3,
                aa.aa_tec1,
                aa.aa_tec2,
                aa.aa_tec3,
                aa.aa_nblet,
                aa.aa_npai,
                aa.aa_dtmaj
            FROM AT_AS#ASSURE a
            JOIN AT_AA#ADR_ASS aa
              ON aa.aaas_id = a.as_id
            LEFT JOIN AT_PY#PAYS py
              ON py.py_id = aa.aapy_id
            WHERE a.as_nni = :nir
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"nir": nir})
                    row = cursor.fetchone()
        except Exception as exc:  # noqa: BLE001
            return {"data": None, "error": str(exc)}

        if not row:
            return {"data": None, "error": None}

        data = {
            "ligne1": row[0],
            "ligne2": row[1],
            "ligne3": row[2],
            "ligne4": row[3],
            "code_postal": row[4],
            "ville": row[5],
            "pays": row[6],
            "email": row[7],
            "tel1": row[8],
            "tel2": row[9],
            "tel3": row[10],
            "fax1": row[11],
            "fax2": row[12],
            "fax3": row[13],
            "nb_lettres": row[14],
            "npai": row[15],
            "date_maj": _format_date(row[16]),
        }
        return {"data": data, "error": None}

    def query_assure_ayants_droit(
        self,
        username: str,
        password: str,
        nir: str,
    ) -> Dict[str, object]:
        """Retourne la liste des ayants droit (par NIR)."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                ay.ay_noord,
                ay.ay_nni,
                ay.ay_nompat,
                ay.ay_nomusuel,
                ay.ay_prenoms,
                ay.ay_dtnais
            FROM AT_AS#ASSURE a
            LEFT JOIN AT_AY#AYANT_DROIT ay
              ON ay.ayas_id = a.as_id
            WHERE a.as_nni = :nir
            ORDER BY ay.ay_noord
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"nir": nir})
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append(
                {
                    "rang": row[0],
                    "nni": row[1],
                    "nom": row[2],
                    "nom_usuel": row[3],
                    "prenoms": row[4],
                    "date_naissance": _format_date(row[5]),
                }
            )
        return {"data": data, "error": None}

    def query_ayant_droit_identification(
        self,
        username: str,
        password: str,
        assure_nir: str,
        ayant_nir: str,
    ) -> Dict[str, object]:
        """Retourne l'identification d'un ayant droit (par NIR assure + NIR ayant droit)."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": None, "error": f"Module oracledb indisponible : {exc}"}

        sql_primary = """
            SELECT
                a.as_nni,
                ay.ay_noord,
                ay.ay_nni,
                ay.ay_nompat,
                ay.ay_nomusuel,
                ay.ay_prenoms,
                ay.ay_dtnais,
                ay.ay_cdsexe,
                ay.ay_dtmaj,
                ay.ay_codcomnais,
                ay.ay_naiscom,
                ay.ay_rgnais,
                pn.py_lib,
                ay.ay_dtdeces,
                ay.ay_dtrniam,
                ay.ay_dtinscr,
                tn.tn_lib,
                pnational.py_lib,
                ay.ay_id
            FROM AT_AS#ASSURE a
            LEFT JOIN AT_AY#AYANT_DROIT ay
              ON ay.ayas_id = a.as_id
            LEFT JOIN AT_PY#PAYS pn
              ON ay.aypy_nais_id = pn.py_id
            LEFT JOIN AT_PY#PAYS pnational
              ON ay.aypy_natpays1_id = pnational.py_id
            LEFT JOIN at_tn#type_national tn
              ON ay.aytn_id = tn.tn_id
            WHERE a.as_nni = :assure_nir
              AND ay.ay_nni = :ayant_nir
        """

        sql_fallback = sql_primary.replace("ay.aypy_natpays1_id", "ay.aypy_nais_id")

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    try:
                        cursor.execute(sql_primary, {"assure_nir": assure_nir, "ayant_nir": ayant_nir})
                        row = cursor.fetchone()
                    except Exception as exc:
                        # Some schemas don't have AYPY_NATPAYS1_ID; fallback to the query provided by the user.
                        if "ORA-00904" in str(exc) and "AYPY_NATPAYS1_ID" in str(exc).upper():
                            cursor.execute(sql_fallback, {"assure_nir": assure_nir, "ayant_nir": ayant_nir})
                            row = cursor.fetchone()
                        else:
                            raise
        except Exception as exc:  # noqa: BLE001
            return {"data": None, "error": str(exc)}

        if not row:
            return {"data": None, "error": None}

        data = {
            "nni_assure": row[0],
            "rang_beneficiaire": row[1],
            "nni_ayant_droit": row[2],
            "nom_ayant_droit": row[3],
            "nom_usuel_ayant_droit": row[4],
            "prenoms_ayant_droit": row[5],
            "date_naissance_ayant_droit": _format_date(row[6]),
            "sexe_ayant_droit": row[7],
            "date_maj": _format_date(row[8]),
            "code_commune_naissance": row[9],
            "commune_de_naissance": row[10],
            "rang_naissance": row[11],
            "pays_de_naissance": row[12],
            "date_deces": _format_date(row[13]),
            "date_certif_rniam": _format_date(row[14]),
            "date_inscription": _format_date(row[15]),
            "type_nationalite": row[16],
            "pays_nationalite": row[17],
            "ayant_droit_id": row[18],
        }
        return {"data": data, "error": None}

    def query_ayant_droit_jod_history(
        self,
        username: str,
        password: str,
        assure_nir: str,
        ayant_nir: str,
    ) -> Dict[str, object]:
        """Retourne l'historique JOD d'un ayant droit (via AY_ID)."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        # Resolve ay_id from NIRs.
        id_resp = self.query_ayant_droit_identification(username, password, assure_nir, ayant_nir)
        if id_resp.get("error"):
            return {"data": [], "error": id_resp["error"]}
        ad = id_resp.get("data") or {}
        ay_id = ad.get("ayant_droit_id")
        if not ay_id:
            return {"data": [], "error": None}

        sql = """
            SELECT
                j.jo_lib,
                jod.jy_dtdjod,
                jod.jy_dtfjod,
                jod.jy_dtmaj
            FROM AT_JY#JOD_AYT jod
            LEFT JOIN AT_JO#COD_JOD j
              ON jod.jyjo_id = j.jo_id
            WHERE jod.jyay_id = :ay_id
            ORDER BY jod.jy_dtdjod DESC
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"ay_id": ay_id})
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append(
                {
                    "situation_jod": row[0],
                    "date_debut": _format_date(row[1]),
                    "date_fin": _format_date(row[2]),
                    "date_maj": _format_date(row[3]),
                }
            )
        return {"data": data, "error": None}

    def query_assure_arpege_summary(
        self,
        username: str,
        password: str,
        nir: str,
    ) -> Dict[str, object]:
        """Retourne la synthese ARPEGE d'un assure (par NIR)."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": None, "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                a.ccg_tvalav36,
                a.ccg_tvalav79,
                a.ccg_tval7997,
                a.ccg_tmajo,
                a.ccg_dtmaj
            FROM AT_CCG#CAR_CAV_GEN a
            LEFT JOIN AT_AS#ASSURE b
              ON a.ccgas_id = b.as_id
            WHERE b.as_nni = :nir
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"nir": nir})
                    row = cursor.fetchone()
        except Exception as exc:  # noqa: BLE001
            return {"data": None, "error": str(exc)}

        if not row:
            return {"data": None, "error": None}

        data = {
            "tval_av36": row[0],
            "tval_36_79": row[1],
            "tval_79_97": row[2],
            "tmaj": row[3],
            "date_maj": _format_date(row[4]),
        }
        return {"data": data, "error": None}

    def query_assure_arpege_detail(
        self,
        username: str,
        password: str,
        nir: str,
    ) -> Dict[str, object]:
        """Retourne le detail ARPEGE par annee (par NIR)."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                cca.cca_annee,
                cca.cca_trival,
                cca.cca_tricot,
                cca.cca_triass,
                cca.cca_jan,
                cca.cca_fev,
                cca.cca_mar,
                cca.cca_avr,
                cca.cca_mai,
                cca.cca_jui,
                cca.cca_jul,
                cca.cca_aou,
                cca.cca_sep,
                cca.cca_oct,
                cca.cca_nov,
                cca.cca_dec,
                cca.cca_dtmaj,
                cca.cca_trirach
            FROM AT_CCA#CAR_CAV_COT cca
            JOIN AT_AS#ASSURE a
              ON a.as_id = cca.ccaas_id
            WHERE a.as_nni = :nir
            ORDER BY cca.cca_annee
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"nir": nir})
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append(
                {
                    "annee": row[0],
                    "trival": row[1],
                    "tricot": row[2],
                    "triass": row[3],
                    "jan": row[4],
                    "fev": row[5],
                    "mar": row[6],
                    "avr": row[7],
                    "mai": row[8],
                    "jui": row[9],
                    "jul": row[10],
                    "aou": row[11],
                    "sep": row[12],
                    "oct": row[13],
                    "nov": row[14],
                    "dec": row[15],
                    "date_maj": _format_date(row[16]),
                    "trirach": row[17],
                }
            )
        return {"data": data, "error": None}

    def query_collectivites(
        self,
        username: str,
        password: str,
        numero: str,
        denom1: str,
        code_postal: str,
    ) -> Dict[str, object]:
        """Recherche des collectivites par numero / denom1 / code postal (prefixes, insensible à la casse)."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        params = {
            "num_pattern": f"{numero}%" if numero else None,
            "denom_pattern": f"{denom1}%" if denom1 else None,
            "cp_pattern": f"{code_postal}%" if code_postal else None,
        }

        sql = """
            SELECT
                cl.cl_id,
                cl.cl_denom1,
                cl.cl_adrcp,
                cl.cl_adrvil
            FROM AT_CL#COLLECTIVITE cl
            WHERE (:num_pattern IS NULL OR UPPER(cl.cl_id) LIKE UPPER(:num_pattern))
              AND (
                :denom_pattern IS NULL
                OR UPPER(cl.cl_denom1) LIKE UPPER(:denom_pattern)
                OR UPPER(cl.cl_adrvil) LIKE UPPER(:denom_pattern)
              )
              AND (:cp_pattern IS NULL OR UPPER(cl.cl_adrcp) LIKE UPPER(:cp_pattern))
            ORDER BY cl.cl_id
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, params)
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append(
                {
                    "numero": row[0],
                    "denom1": row[1],
                    "denom2": None,
                    "cp": row[2],
                    "ville": row[3],
                }
            )
        return {"data": data, "error": None}

    def query_collectivite_adresse(
        self,
        username: str,
        password: str,
        collect_id: str,
    ) -> Dict[str, object]:
        """Récupère les informations d'adresse d'une collectivité par identifiant."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": None, "error": f"Module oracledb indisponible : {exc}"}

        if not collect_id:
            return {"data": None, "error": "Identifiant collectivité manquant."}

        sql = """
            SELECT
                cl.cl_id,
                cl.cl_denom1,
                cl.cl_denom2,
                cl.cl_adrcp,
                cl.cl_adr1,
                cl.cl_adr2,
                cl.cl_adr3,
                cl.cl_adr4,
                cl.cl_adrvil,
                p.py_lib,
                cl.cl_email,
                cl.cl_tel1,
                cl.cl_tec1,
                cl.cl_npai,
                cl.cl_dtadr
            FROM AT_CL#COLLECTIVITE cl
            LEFT JOIN at_py#pays p on cl.clpy_id = p.py_id
            WHERE cl.cl_id = :collect_id
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"collect_id": collect_id})
                    row = cursor.fetchone()
        except Exception as exc:  # noqa: BLE001
            return {"data": None, "error": str(exc)}

        if not row:
            return {"data": None, "error": None}

        data = {
            "numero": row[0],
            "denom1": row[1],
            "denom2": row[2],
            "cp": row[3],
            "adr1": row[4],
            "adr2": row[5],
            "adr3": row[6],
            "adr4": row[7],
            "ville": row[8],
            "pays": row[9],
            "email": row[10],
            "tel": row[11],
            "fax": row[12],
            "npai": row[13],
            "date_maj": _format_date(row[14]) if hasattr(row[14], "strftime") else row[14],
        }
        return {"data": data, "error": None}

    def query_collectivite_identification(
        self,
        username: str,
        password: str,
        collect_id: str,
    ) -> Dict[str, object]:
        """Recupere les informations d'identification d'une collectivite par identifiant."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": None, "error": f"Module oracledb indisponible : {exc}"}

        if not collect_id:
            return {"data": None, "error": "Identifiant collectivite manquant."}

        sql = """
            SELECT
                cl.cl_id,
                cl.cl_denom1,
                cl.cl_denom2,
                cu.cu_lib,
                cl.cl_dtdec,
                cl.cl_dtcrjo,
                mv.vm_lib,
                cl.cl_dtrec,
                cl.cl_dtmaj,
                cl.cl_nblet
            FROM AT_CL#COLLECTIVITE cl
            LEFT JOIN AT_CU#Culte cu on cl.clcu_id = cu.cu_id
            LEFT JOIN at_vm#mode_vie mv on cl.clvm_id = mv.vm_id
            WHERE cl.cl_id = :collect_id
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"collect_id": collect_id})
                    row = cursor.fetchone()
        except Exception as exc:  # noqa: BLE001
            return {"data": None, "error": str(exc)}

        if not row:
            return {"data": None, "error": None}

        data = {
            "numero": row[0],
            "denom1": row[1],
            "denom2": row[2],
            "culte": row[3],
            "date_adhesion": _format_date(row[4]) if hasattr(row[4], "strftime") else row[4],
            "date_journal": _format_date(row[5]) if hasattr(row[5], "strftime") else row[5],
            "mode_vie": row[6],
            "recult": _format_date(row[7]) if hasattr(row[7], "strftime") else row[7],
            "date_maj": _format_date(row[8]) if hasattr(row[8], "strftime") else row[8],
            "nb_lettres": row[9],
        }
        return {"data": data, "error": None}

    def query_collectivite_situations(
        self,
        username: str,
        password: str,
        collect_id: str,
    ) -> Dict[str, object]:
        """Recupere l'historique des situations pour une collectivite."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        if not collect_id:
            return {"data": [], "error": "Identifiant collectivite manquant."}

        sql = """
            SELECT
                his.hc_dtdeb,
                cs.cs_lib,
                CASE
                    WHEN his.hccs_id = 'U' THEN fc.fccl2_id
                    ELSE ''
                END AS coll_accueil,
                his.hc_dtmaj
            FROM at_hc#his_sit_col his
            LEFT JOIN at_cs#lib_sit_col cs on his.hccs_id = cs.cs_id
            LEFT JOIN at_fc#fusion_col fc on his.HCCL_ID = fc.fccl1_id
            WHERE his.HCCL_ID = :collect_id
            ORDER BY his.hc_dtmaj desc
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"collect_id": collect_id})
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append(
                {
                    "date_effet": _format_date(row[0]) if hasattr(row[0], "strftime") else row[0],
                    "libelle": row[1],
                    "coll_accueil": row[2],
                    "date_maj": _format_date(row[3]) if hasattr(row[3], "strftime") else row[3],
                }
            )
        return {"data": data, "error": None}

    def query_collectivite_fusions(
        self,
        username: str,
        password: str,
        collect_id: str,
    ) -> Dict[str, object]:
        """Recupere l'historique des collectivites reprises suite a fusion."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        if not collect_id:
            return {"data": [], "error": "Identifiant collectivite manquant."}

        sql = """
            SELECT
                a.fc_dteff,
                a.fccl1_id,
                a.fc_dtmaj
            FROM at_fc#fusion_col a
            WHERE a.fccl2_id = :collect_id
            ORDER BY a.fc_dtmaj desc
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"collect_id": collect_id})
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append(
                {
                    "date_effet": _format_date(row[0]) if hasattr(row[0], "strftime") else row[0],
                    "coll_transf": row[1],
                    "date_maj": _format_date(row[2]) if hasattr(row[2], "strftime") else row[2],
                }
            )
        return {"data": data, "error": None}

    def query_collectivite_responsable(
        self,
        username: str,
        password: str,
        collect_id: str,
        role_id: int,
    ) -> Dict[str, object]:
        """Recupere les informations d'un responsable (maladie/vieillesse) pour une collectivite."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": None, "error": f"Module oracledb indisponible : {exc}"}

        if not collect_id:
            return {"data": None, "error": "Identifiant collectivite manquant."}

        sql = """
            SELECT
                responsable.rs_nomrespo,
                responsable.rs_adr1,
                responsable.rs_adr2,
                responsable.rs_adr3,
                responsable.rs_adr4,
                responsable.rs_adrcp,
                responsable.rs_adrvil,
                p_adresse.py_lib,
                responsable.rs_email,
                responsable.rs_tel1,
                responsable.rs_tec1,
                responsable.rs_dtmaj
            FROM AT_CL#COLLECTIVITE a
            LEFT JOIN AT_RS#RESPONSABLE responsable
                ON a.cl_id = responsable.rscl_id
                AND responsable.rsst_id = :role_id
            LEFT JOIN at_py#pays p_adresse
                ON responsable.rspy_id = p_adresse.py_id
            WHERE a.cl_id = :collect_id
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"collect_id": collect_id, "role_id": role_id})
                    row = cursor.fetchone()
        except Exception as exc:  # noqa: BLE001
            return {"data": None, "error": str(exc)}

        if not row:
            return {"data": None, "error": None}

        data = {
            "nom": row[0],
            "adr1": row[1],
            "adr2": row[2],
            "adr3": row[3],
            "adr4": row[4],
            "cp": row[5],
            "ville": row[6],
            "pays": row[7],
            "email": row[8],
            "tel": row[9],
            "fax": row[10],
            "date_maj": _format_date(row[11]) if hasattr(row[11], "strftime") else row[11],
        }
        return {"data": data, "error": None}

    def query_collectivite_assures(
        self,
        username: str,
        password: str,
        collect_id: str,
        filter_type: str,
        order_by: str,
    ) -> Dict[str, object]:
        """Recupere la liste des assures d'une collectivite."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        if not collect_id:
            return {"data": [], "error": "Identifiant collectivite manquant."}

        filter_clause = ""
        if filter_type == "maladie":
            filter_clause = """
              AND sm.smsa_id IS NOT NULL
              AND (sm.sm_dteffin = TO_DATE('31-12-3999', 'DD-MM-YYYY') OR sm.sm_dteffin IS NULL)
            """
        elif filter_type == "vieillesse":
            filter_clause = """
              AND sv.svsa_id IS NOT NULL
              AND (sv.sv_dteffin = TO_DATE('31-12-3999', 'DD-MM-YYYY') OR sv.sv_dteffin IS NULL)
            """
        elif filter_type == "adresse":
            filter_clause = " AND b.as_typadr = 2"

        order_field = "b.as_nompat" if order_by != "nni" else "b.as_nni"

        sql = f"""
            WITH
            sm_last AS (
              SELECT *
              FROM (
                SELECT
                  c.*,
                  ROW_NUMBER() OVER (
                    PARTITION BY c.smas_id
                    ORDER BY
                      CASE WHEN c.sm_dteffin IS NULL THEN 1 ELSE 0 END DESC,
                      c.sm_dteffin DESC,
                      c.sm_dtefdeb DESC
                  ) rn
                FROM AT_SM#ASS_SIT_CAM c
              )
              WHERE rn = 1
            ),
            sv_last AS (
              SELECT *
              FROM (
                SELECT
                  d.*,
                  ROW_NUMBER() OVER (
                    PARTITION BY d.svas_id
                    ORDER BY
                      CASE WHEN d.sv_dteffin IS NULL THEN 1 ELSE 0 END DESC,
                      d.sv_dteffin DESC,
                      d.sv_dtefdeb DESC
                  ) rn
                FROM AT_SV#ASS_SIT_VIC d
              )
              WHERE rn = 1
            )
            SELECT DISTINCT
              b.as_nompat,
              b.as_prenoms,
              b.as_nni,
              sm.smsa_id,
              sm.sm_dtefdeb,
              sm.sm_dteffin,
              a.accl_id,
              sv.svsa_id,
              sv.sv_dtefdeb,
              sv.sv_dteffin,
              a.accl_id,
              CASE
                WHEN b.as_typadr = 1 THEN 'Assuré'
                ELSE 'Collectivité'
              END AS type_adresse
            FROM AT_AC#ASS_COL a
            LEFT JOIN AT_AS#ASSURE b
              ON a.acas_id = b.as_id
            LEFT JOIN sm_last sm
              ON b.as_id = sm.smas_id
            LEFT JOIN sv_last sv
              ON b.as_id = sv.svas_id
            WHERE a.accl_id = :collect_id
              AND a.ac_dteffin = TO_DATE('31-12-3999', 'DD-MM-YYYY')
              {filter_clause}
            ORDER BY {order_field}
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"collect_id": collect_id})
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append(
                {
                    "nom": row[0],
                    "prenoms": row[1],
                    "nni": row[2],
                    "code_maladie": row[3],
                    "date_effet_maladie": _format_date(row[4]) if hasattr(row[4], "strftime") else row[4],
                    "date_fin_maladie": _format_date(row[5]) if hasattr(row[5], "strftime") else row[5],
                    "num_coll_maladie": row[6],
                    "code_vieillesse": row[7],
                    "date_effet_vieillesse": _format_date(row[8]) if hasattr(row[8], "strftime") else row[8],
                    "date_fin_vieillesse": _format_date(row[9]) if hasattr(row[9], "strftime") else row[9],
                    "num_coll_vieillesse": row[10],
                    "type_adresse": row[11],
                }
            )
        return {"data": data, "error": None}

    def query_collectivite_communautes(
        self,
        username: str,
        password: str,
        collect_id: str,
    ) -> Dict[str, object]:
        """Recupere la liste des communautes pour une collectivite."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        if not collect_id:
            return {"data": [], "error": "Identifiant collectivite manquant."}

        sql = """
            SELECT
                cl.cl_id,
                ROW_NUMBER() OVER (ORDER BY co.co_denom1) AS rang,
                co.co_denom1,
                co.co_denom2,
                co.co_adrcp,
                co.co_adrvil
            FROM AT_CO#COMMUNAUTE co
            LEFT JOIN at_cl#collectivite cl
                ON cl.cl_id = co.cocl_id
            WHERE cl.cl_id = :collect_id
            ORDER BY co.co_denom1
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"collect_id": collect_id})
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append(
                {
                    "numero": row[0],
                    "rang": row[1],
                    "denom1": row[2],
                    "denom2": row[3],
                    "cp": row[4],
                    "ville": row[5],
                }
            )
        return {"data": data, "error": None}

    def query_collectivite_referent_maladie(
        self,
        username: str,
        password: str,
        collect_id: str,
    ) -> Dict[str, object]:
        """Recupere le referent maladie pour une collectivite."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": None, "error": f"Module oracledb indisponible : {exc}"}

        if not collect_id:
            return {"data": None, "error": "Identifiant collectivite manquant."}

        sql = """
            SELECT
                resp.rmc_nom,
                resp.rmc_prenom,
                resp.rmc_numvoie,
                resp.rmc_libvoie,
                resp.rmc_comad,
                resp.rmc_codpos,
                resp.rmc_burdis,
                p.py_lib,
                resp.rmc_email,
                resp.rmc_tel1,
                resp.rmc_tec1,
                resp.rmc_dtmaj
            FROM AT_RMC#REF_MAL_COLL resp
            LEFT JOIN at_cl#collectivite cl on resp.rmccl_id = cl.cl_id
            LEFT JOIN at_py#pays p on resp.rmcpy_id = p.py_id
            WHERE cl.cl_id = :collect_id
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql, {"collect_id": collect_id})
                    row = cursor.fetchone()
        except Exception as exc:  # noqa: BLE001
            return {"data": None, "error": str(exc)}

        if not row:
            return {"data": None, "error": None}

        data = {
            "nom": row[0],
            "prenom": row[1],
            "num_voie": row[2],
            "lib_voie": row[3],
            "complement": row[4],
            "cp": row[5],
            "burdis": row[6],
            "pays": row[7],
            "email": row[8],
            "tel": row[9],
            "fax": row[10],
            "date_maj": _format_date(row[11]) if hasattr(row[11], "strftime") else row[11],
        }
        return {"data": data, "error": None}

    def query_referentiel_civilites(self, username: str, password: str) -> Dict[str, object]:
        """Retourne la liste des civilites."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                cv.cv_id,
                cv.cv_lib
            FROM AT_CV#CIVILITE cv
            ORDER BY cv.cv_id
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql)
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append({"code": row[0], "libelle": row[1]})
        return {"data": data, "error": None}

    def query_referentiel_cultes(self, username: str, password: str) -> Dict[str, object]:
        """Retourne la liste des cultes."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                cu.cu_id,
                cu.cu_lib
            FROM AT_CU#CULTE cu
            ORDER BY cu.cu_id
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql)
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append({"code": row[0], "libelle": row[1]})
        return {"data": data, "error": None}

    def query_referentiel_complements_num_voie(self, username: str, password: str) -> Dict[str, object]:
        """Retourne la liste des complements numero de voie."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                cc.cc_id,
                cc.cc_lib
            FROM AT_CC#COM_NUM_VOI cc
            ORDER BY cc.cc_id
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql)
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append({"code": row[0], "libelle": row[1]})
        return {"data": data, "error": None}

    def query_referentiel_jod(self, username: str, password: str) -> Dict[str, object]:
        """Retourne la liste des justificatifs d'ouverture des droits (J.O.D)."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                jo.jo_id,
                jo.jo_lib
            FROM AT_JO#COD_JOD jo
            ORDER BY jo.jo_id
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql)
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append({"code": row[0], "libelle": row[1]})
        return {"data": data, "error": None}

    def query_referentiel_mode_vie(self, username: str, password: str) -> Dict[str, object]:
        """Retourne la liste des modes de vie."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                vm.vm_id,
                vm.vm_lib
            FROM AT_VM#MODE_VIE vm
            ORDER BY vm.vm_id
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql)
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append({"code": row[0], "libelle": row[1]})
        return {"data": data, "error": None}

    def query_referentiel_nature_situation(self, username: str, password: str) -> Dict[str, object]:
        """Retourne la liste des natures de situation."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                ns.ns_id,
                ns.ns_lib
            FROM AT_NS#LIB_NAT_SIT ns
            ORDER BY ns.ns_id
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql)
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append({"code": row[0], "libelle": row[1]})
        return {"data": data, "error": None}

    def query_referentiel_pays(self, username: str, password: str) -> Dict[str, object]:
        """Retourne la liste des pays."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                py.py_id,
                py.py_lib
            FROM AT_PY#PAYS py
            ORDER BY py.py_lib
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql)
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append({"code": row[0], "libelle": row[1]})
        return {"data": data, "error": None}

    def query_referentiel_situations(self, username: str, password: str) -> Dict[str, object]:
        """Retourne la liste des situations assure."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                st.st_lib,
                sa.sa_id,
                sa.sa_lib
            FROM AT_SA#LIB_SIT_ASS sa
            LEFT JOIN at_st#societe st ON sa.sast_id = st.st_id
            ORDER BY st.st_id ASC
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql)
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append({"categorie": row[0], "code": row[1], "libelle": row[2]})
        return {"data": data, "error": None}

    def query_referentiel_situations_collectivite(self, username: str, password: str) -> Dict[str, object]:
        """Retourne la liste des situations collectivite."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                cs.cs_id,
                cs.cs_lib
            FROM AT_CS#LIB_SIT_COL cs
            ORDER BY cs.cs_id
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql)
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append({"code": row[0], "libelle": row[1]})
        return {"data": data, "error": None}

    def query_referentiel_type_nationalite(self, username: str, password: str) -> Dict[str, object]:
        """Retourne la liste des types de nationalite."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                tn.tn_id,
                tn.tn_lib
            FROM AT_TN#TYPE_NATIONAL tn
            ORDER BY tn.tn_id
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql)
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append({"code": row[0], "libelle": row[1]})
        return {"data": data, "error": None}

    def query_referentiel_type_voie(self, username: str, password: str) -> Dict[str, object]:
        """Retourne la liste des types de voie."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                cn.cn_id,
                cn.cn_lib
            FROM AT_CN#TYP_VOIE cn
            ORDER BY cn.cn_id
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql)
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append({"code": row[0], "libelle": row[1]})
        return {"data": data, "error": None}

    def query_referentiel_societe(self, username: str, password: str) -> Dict[str, object]:
        """Retourne la liste des societes."""
        try:
            import oracledb
        except Exception as exc:
            return {"data": [], "error": f"Module oracledb indisponible : {exc}"}

        sql = """
            SELECT
                st.st_id,
                st.st_lib
            FROM AT_ST#SOCIETE st
            ORDER BY st.st_id
        """

        try:
            with oracledb.connect(user=username, password=password, dsn=self.dsn) as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql)
                    rows = cursor.fetchall()
        except Exception as exc:  # noqa: BLE001
            return {"data": [], "error": str(exc)}

        data: List[Dict[str, object]] = []
        for row in rows:
            data.append({"code": row[0], "libelle": row[1]})
        return {"data": data, "error": None}


def build_assure_workbook(
    info: Dict[str, object],
    situation_maladie: Optional[Dict[str, object]],
    situation_vieillesse: Optional[Dict[str, object]],
    historique_maladie: List[Dict[str, object]],
    collectivites_maladie: List[Dict[str, object]],
    historique_vieillesse: List[Dict[str, object]],
    collectivites_vieillesse: List[Dict[str, object]],
    adresse: Optional[Dict[str, object]],
    ayants_droit: List[Dict[str, object]],
    arpege_summary: Optional[Dict[str, object]],
    arpege_details: List[Dict[str, object]],
    include_arpege: bool = True,
):
    """Construit un classeur Excel pour un assure (par onglet)."""
    from openpyxl import Workbook

    def _safe(value: object) -> str:
        if value is None or value == "":
            return "-"
        return str(value)

    def _append_section(ws, title: str, rows: List[Tuple[str, object]]) -> None:
        ws.append([title])
        ws.append(["Champ", "Valeur"])
        for label, val in rows:
            ws.append([label, _safe(val)])
        ws.append([])

    def _append_table(ws, title: str, headers: List[str], rows: List[List[object]]) -> None:
        ws.append([title])
        ws.append(headers)
        if not rows:
            ws.append(["-"] * len(headers))
        else:
            for row in rows:
                ws.append([_safe(v) for v in row])
        ws.append([])

    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "Infos personnels"
    _append_section(
        ws_info,
        "Identification",
        [
            ("NIR", info.get("nir")),
            ("Nom usage", info.get("nom_usage")),
            ("Prenoms", info.get("prenom_usage")),
            ("Date naissance", info.get("date_naissance")),
            ("Civilite", info.get("civilite")),
            ("Nom usuel", info.get("nom_usuel")),
            ("Date certification RNIAM", info.get("date_certif_rniam")),
            ("Collectivite maladie", info.get("collectivite_maladie")),
            ("Collectivite vieillesse", info.get("collectivite_vieillesse")),
            ("Code postal naissance", info.get("code_postal")),
            ("Commune naissance", info.get("commune")),
            ("Pays naissance", info.get("pays_naissance")),
            ("Type nationalite", info.get("type_nationalite")),
            ("Pays nationalite 1", info.get("pays1")),
            ("Pays nationalite 2", info.get("pays2")),
            ("Date naturalisation", info.get("date_naturalisation")),
            ("Entree en vie religieuse", info.get("date_entree_vie_religieuse")),
            ("Cessation vie religieuse", info.get("date_cessation_vie_religieuse")),
            ("Date fin visa", info.get("date_fin_visa")),
            ("Date MAJ", info.get("date_maj")),
        ],
    )

    ws_situ = wb.create_sheet("Situations")
    _append_table(
        ws_situ,
        "Situation maladie (en cours)",
        [
            "Code situation",
            "Nature 1",
            "Nature 2",
            "Date nature 2",
            "Date condition",
            "Date declaration",
            "Date effet",
            "Date MAJ situation",
            "Date MAJ",
        ],
        [
            [
                (situation_maladie or {}).get("code_situation"),
                (situation_maladie or {}).get("code_nature1"),
                (situation_maladie or {}).get("code_nature2"),
                (situation_maladie or {}).get("date_nature2"),
                (situation_maladie or {}).get("date_conditions"),
                (situation_maladie or {}).get("date_declaration"),
                (situation_maladie or {}).get("date_effet"),
                (situation_maladie or {}).get("date_maj_situation"),
                (situation_maladie or {}).get("date_maj"),
            ]
            if situation_maladie
            else []
        ],
    )
    _append_table(
        ws_situ,
        "Situation vieillesse (en cours)",
        [
            "Code situation",
            "Nature 1",
            "Nature 2",
            "Date nature 2",
            "Date condition",
            "Date declaration",
            "Date effet",
            "Date MAJ situation",
            "Date MAJ",
        ],
        [
            [
                (situation_vieillesse or {}).get("code_situation"),
                (situation_vieillesse or {}).get("code_nature1"),
                (situation_vieillesse or {}).get("code_nature2"),
                (situation_vieillesse or {}).get("date_nature2"),
                (situation_vieillesse or {}).get("date_conditions"),
                (situation_vieillesse or {}).get("date_declaration"),
                (situation_vieillesse or {}).get("date_effet"),
                (situation_vieillesse or {}).get("date_maj_situation"),
                (situation_vieillesse or {}).get("date_maj"),
            ]
            if situation_vieillesse
            else []
        ],
    )

    ws_mal = wb.create_sheet("Maladie")
    _append_table(
        ws_mal,
        "Historique des situations maladie",
        [
            "Code situation",
            "Code nature 1",
            "Code nature 2",
            "Date nature 2",
            "Date condition",
            "Date declaration",
            "Date effet",
            "Date MAJ situation",
            "Date MAJ",
        ],
        [
            [
                r.get("code_situation"),
                r.get("code_nature1"),
                r.get("code_nature2"),
                r.get("date_nature2"),
                r.get("date_conditions"),
                r.get("date_declaration"),
                r.get("date_effet"),
                r.get("date_maj_situation"),
                r.get("date_maj"),
            ]
            for r in (historique_maladie or [])
        ],
    )
    _append_table(
        ws_mal,
        "Historique des collectivites maladie",
        ["No collectivite", "Date effet", "Date MAJ", "Etat actuel"],
        [
            [
                r.get("collectivite"),
                r.get("date_effet"),
                r.get("date_maj"),
                r.get("etat_actuel"),
            ]
            for r in (collectivites_maladie or [])
        ],
    )

    ws_vie = wb.create_sheet("Vieillesse")
    _append_table(
        ws_vie,
        "Historique des situations vieillesse",
        [
            "Code situation",
            "Code nature 1",
            "Code nature 2",
            "Date nature 2",
            "Date condition",
            "Date declaration",
            "Date effet",
            "Date MAJ situation",
            "Date MAJ",
        ],
        [
            [
                r.get("code_situation"),
                r.get("code_nature1"),
                r.get("code_nature2"),
                r.get("date_nature2"),
                r.get("date_conditions"),
                r.get("date_declaration"),
                r.get("date_effet"),
                r.get("date_maj_situation"),
                r.get("date_maj"),
            ]
            for r in (historique_vieillesse or [])
        ],
    )
    _append_table(
        ws_vie,
        "Historique des collectivites vieillesse",
        ["No collectivite", "Date effet", "Date MAJ", "Etat actuel"],
        [
            [
                r.get("collectivite"),
                r.get("date_effet"),
                r.get("date_maj"),
                r.get("etat_actuel"),
            ]
            for r in (collectivites_vieillesse or [])
        ],
    )

    ws_addr = wb.create_sheet("Adresse")
    adresse = adresse or {}
    _append_section(
        ws_addr,
        "Adresse",
        [
            ("Ligne 1", adresse.get("ligne1")),
            ("Ligne 2", adresse.get("ligne2")),
            ("Ligne 3", adresse.get("ligne3")),
            ("Ligne 4", adresse.get("ligne4")),
            ("Code postal", adresse.get("code_postal")),
            ("Ville", adresse.get("ville")),
            ("Pays", adresse.get("pays")),
        ],
    )
    _append_section(
        ws_addr,
        "Coordonnees",
        [
            ("Email", adresse.get("email")),
            ("Telephone 1", adresse.get("tel1")),
            ("Telephone 2", adresse.get("tel2")),
            ("Telephone 3", adresse.get("tel3")),
            ("Telecopie 1", adresse.get("fax1")),
            ("Telecopie 2", adresse.get("fax2")),
            ("Telecopie 3", adresse.get("fax3")),
            ("Nombre lettres info", adresse.get("nb_lettres")),
            ("NPAI", adresse.get("npai")),
            ("Date MAJ", adresse.get("date_maj")),
        ],
    )

    ws_ay = wb.create_sheet("Ayants droit")
    ws_ay.append(["Rang", "Nom", "Nom usuel", "Prenoms", "Date naissance"])
    if not ayants_droit:
        ws_ay.append(["-"] * 5)
    else:
        for r in ayants_droit:
            ws_ay.append(
                [
                    _safe(r.get("rang")),
                    _safe(r.get("nom")),
                    _safe(r.get("nom_usuel")),
                    _safe(r.get("prenoms")),
                    _safe(r.get("date_naissance")),
                ]
            )

    if include_arpege:
        ws_arp = wb.create_sheet("ARPEGE")
        _append_section(
            ws_arp,
            "Synthese ARPEGE",
            [
                ("TVAL < 1936", (arpege_summary or {}).get("tval_av36")),
                ("TVAL 1936-1979", (arpege_summary or {}).get("tval_36_79")),
                ("TVAL 1979-1997", (arpege_summary or {}).get("tval_79_97")),
                ("Trimestres majores", (arpege_summary or {}).get("tmaj")),
                ("Date MAJ", (arpege_summary or {}).get("date_maj")),
            ],
        )
        _append_table(
            ws_arp,
            "Detail par annee",
            [
                "Annee",
                "TriVal",
                "TriCot",
                "TriAss",
                "Jan",
                "Fev",
                "Mar",
                "Avr",
                "Mai",
                "Jui",
                "Jul",
                "Aou",
                "Sep",
                "Oct",
                "Nov",
                "Dec",
                "Date MAJ",
                "TriRach",
            ],
            [
                [
                    r.get("annee"),
                    r.get("trival"),
                    r.get("tricot"),
                    r.get("triass"),
                    r.get("jan"),
                    r.get("fev"),
                    r.get("mar"),
                    r.get("avr"),
                    r.get("mai"),
                    r.get("jui"),
                    r.get("jul"),
                    r.get("aou"),
                    r.get("sep"),
                    r.get("oct"),
                    r.get("nov"),
                    r.get("dec"),
                    r.get("date_maj"),
                    r.get("trirach"),
                ]
                for r in (arpege_details or [])
            ],
        )

    return wb
def build_collectivite_workbook(
    collect_id: str,
    identification: Optional[Dict[str, object]],
    adresse: Optional[Dict[str, object]],
    responsable_maladie: Optional[Dict[str, object]],
    responsable_vieillesse: Optional[Dict[str, object]],
    assures: List[Dict[str, object]],
    communautes: List[Dict[str, object]],
    referent: Optional[Dict[str, object]],
    situations: List[Dict[str, object]],
    fusions: List[Dict[str, object]],
):
    """Construit un classeur Excel pour une collectivité."""
    from openpyxl import Workbook

    def _safe(value: object) -> str:
        if value is None or value == "":
            return "-"
        return str(value)

    def _append_section(ws, title: str, rows: List[Tuple[str, object]]) -> None:
        ws.append([title])
        ws.append(["Champ", "Valeur"])
        for label, val in rows:
            ws.append([label, _safe(val)])
        ws.append([])

    wb = Workbook()

    ws_ident = wb.active
    ws_ident.title = "Identification"
    _append_section(
        ws_ident,
        "Identification",
        [
            ("Numéro", collect_id),
            ("Dénomination 1", (identification or {}).get("denom1")),
            ("Dénomination 2", (identification or {}).get("denom2")),
            ("Culte", (identification or {}).get("culte")),
            ("Mode de vie", (identification or {}).get("mode_vie")),
            ("1ère adhésion", (identification or {}).get("date_adhesion")),
            ("Création journal officiel", (identification or {}).get("date_journal")),
            ("Reconnaissance cultuelle", (identification or {}).get("recult")),
            ("Date MAJ", (identification or {}).get("date_maj")),
            ("Nb lettres d'informations", (identification or {}).get("nb_lettres")),
        ],
    )

    ws_ident.append(["Historique des situations"])
    ws_ident.append(["Date effet", "Libellé situation", "N° coll. accueil", "Date MAJ"])
    for item in situations or []:
        ws_ident.append(
            [
                _safe(item.get("date_effet")),
                _safe(item.get("libelle")),
                _safe(item.get("coll_accueil")),
                _safe(item.get("date_maj")),
            ]
        )
    ws_ident.append([])
    ws_ident.append(["Collectivités reprises suite à fusion"])
    ws_ident.append(["Date effet", "N° coll. transférée", "Date MAJ"])
    for item in fusions or []:
        ws_ident.append(
            [
                _safe(item.get("date_effet")),
                _safe(item.get("coll_transf")),
                _safe(item.get("date_maj")),
            ]
        )

    ws_addr = wb.create_sheet("Adresse")
    _append_section(
        ws_addr,
        "Adresse",
        [
            ("Ligne 1", (adresse or {}).get("adr1")),
            ("Ligne 2", (adresse or {}).get("adr2")),
            ("Ligne 3", (adresse or {}).get("adr3")),
            ("Ligne 4", (adresse or {}).get("adr4")),
            ("Code postal", (adresse or {}).get("cp")),
            ("Ville", (adresse or {}).get("ville")),
            ("Pays", (adresse or {}).get("pays")),
        ],
    )
    _append_section(
        ws_addr,
        "Coordonnées",
        [
            ("Email", (adresse or {}).get("email")),
            ("Téléphone", (adresse or {}).get("tel")),
            ("Télécopie", (adresse or {}).get("fax")),
            ("NPAI", (adresse or {}).get("npai")),
            ("Date MAJ adresse", (adresse or {}).get("date_maj")),
        ],
    )

    ws_resp = wb.create_sheet("Responsables")
    _append_section(
        ws_resp,
        "Responsable maladie",
        [
            ("Nom du responsable", (responsable_maladie or {}).get("nom")),
            ("Ligne 1", (responsable_maladie or {}).get("adr1")),
            ("Ligne 2", (responsable_maladie or {}).get("adr2")),
            ("Ligne 3", (responsable_maladie or {}).get("adr3")),
            ("Ligne 4", (responsable_maladie or {}).get("adr4")),
            ("Code postal", (responsable_maladie or {}).get("cp")),
            ("Ville", (responsable_maladie or {}).get("ville")),
            ("Pays", (responsable_maladie or {}).get("pays")),
            ("Email", (responsable_maladie or {}).get("email")),
            ("Téléphone", (responsable_maladie or {}).get("tel")),
            ("Télécopie", (responsable_maladie or {}).get("fax")),
            ("Date MAJ", (responsable_maladie or {}).get("date_maj")),
        ],
    )
    _append_section(
        ws_resp,
        "Responsable vieillesse",
        [
            ("Nom du responsable", (responsable_vieillesse or {}).get("nom")),
            ("Ligne 1", (responsable_vieillesse or {}).get("adr1")),
            ("Ligne 2", (responsable_vieillesse or {}).get("adr2")),
            ("Ligne 3", (responsable_vieillesse or {}).get("adr3")),
            ("Ligne 4", (responsable_vieillesse or {}).get("adr4")),
            ("Code postal", (responsable_vieillesse or {}).get("cp")),
            ("Ville", (responsable_vieillesse or {}).get("ville")),
            ("Pays", (responsable_vieillesse or {}).get("pays")),
            ("Email", (responsable_vieillesse or {}).get("email")),
            ("Téléphone", (responsable_vieillesse or {}).get("tel")),
            ("Télécopie", (responsable_vieillesse or {}).get("fax")),
            ("Date MAJ", (responsable_vieillesse or {}).get("date_maj")),
        ],
    )

    ws_assures = wb.create_sheet("Assures")
    ws_assures.append(
        [
            "Nom",
            "Prénoms",
            "NNI",
            "Code sit. mal.",
            "Date effet sit. mal.",
            "Date fin sit. mal.",
            "Num. coll maladie en cours",
            "Code sit. vieil.",
            "Date effet sit. vieil.",
            "Date fin sit. vieil.",
            "Num. coll vieillesse en cours",
            "Type adresse",
        ]
    )
    for item in assures or []:
        ws_assures.append(
            [
                _safe(item.get("nom")),
                _safe(item.get("prenoms")),
                _safe(item.get("nni")),
                _safe(item.get("code_maladie")),
                _safe(item.get("date_effet_maladie")),
                _safe(item.get("date_fin_maladie")),
                _safe(item.get("num_coll_maladie")),
                _safe(item.get("code_vieillesse")),
                _safe(item.get("date_effet_vieillesse")),
                _safe(item.get("date_fin_vieillesse")),
                _safe(item.get("num_coll_vieillesse")),
                _safe(item.get("type_adresse")),
            ]
        )

    ws_commu = wb.create_sheet("Communautes")
    ws_commu.append(["N° Collectivité", "Rang", "Dénomination 1", "Dénomination 2", "Code postal", "Ville"])
    for item in communautes or []:
        ws_commu.append(
            [
                _safe(item.get("numero")),
                _safe(item.get("rang")),
                _safe(item.get("denom1")),
                _safe(item.get("denom2")),
                _safe(item.get("cp")),
                _safe(item.get("ville")),
            ]
        )

    ws_ref = wb.create_sheet("Referent")
    _append_section(
        ws_ref,
        "Référent maladie",
        [
            ("Nom", (referent or {}).get("nom")),
            ("Prénom", (referent or {}).get("prenom")),
            ("Numéro de voie", (referent or {}).get("num_voie")),
            ("Libellé de voie", (referent or {}).get("lib_voie")),
            ("Complément adresse", (referent or {}).get("complement")),
            ("Code postal", (referent or {}).get("cp")),
            ("Bureau distributeur", (referent or {}).get("burdis")),
            ("Pays", (referent or {}).get("pays")),
        ],
    )
    _append_section(
        ws_ref,
        "Coordonnées",
        [
            ("Email", (referent or {}).get("email")),
            ("Téléphone", (referent or {}).get("tel")),
            ("Télécopie", (referent or {}).get("fax")),
            ("Date MAJ", (referent or {}).get("date_maj")),
        ],
    )

    return wb
